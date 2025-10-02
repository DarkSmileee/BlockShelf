"""
Shared helper functions and utilities for views.
Includes type hints for better code clarity and IDE support.
"""

import csv
import gzip
import io
import os
import re
import shutil
import tempfile
import zipfile
from io import BytesIO
from typing import Any, Iterator

import requests
from django.contrib.auth import get_user_model
from django.http import Http404, HttpRequest, JsonResponse
from django.utils.crypto import get_random_string

from ..constants import ELEMENT_ID_MIN_LENGTH, EXTERNAL_API_TIMEOUT
from ..models import InventoryCollab
from ..utils import sanitize_text, sanitize_url

User = get_user_model()


# -------------------------------------------------------------------------------------------------
# JSON Response Helpers
# -------------------------------------------------------------------------------------------------

def json_ok(**payload: Any) -> JsonResponse:
    """Return a successful JSON response with ok=True."""
    data = {"ok": True}
    data.update(payload)
    return JsonResponse(data)


def json_err(msg: str, status: int = 400) -> JsonResponse:
    """Return an error JSON response with ok=False."""
    return JsonResponse({"ok": False, "error": msg}, status=status)


# -------------------------------------------------------------------------------------------------
# File & Path Helpers
# -------------------------------------------------------------------------------------------------

def create_temp_dir() -> str:
    """Create a temporary directory for Rebrickable imports."""
    return tempfile.mkdtemp(prefix="rebrickable_")


def count_csv_rows(path: str) -> int:
    """
    Count data rows in a CSV file (excluding header).
    Supports both .csv and .csv.gz files.
    """
    opener = gzip.open if path.lower().endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        try:
            next(reader)  # Skip header
        except StopIteration:
            return 0
        return sum(1 for _ in reader)


def extract_rebrickable_csvs(zip_path: str, dest_dir: str) -> dict[str, str]:
    """
    Extract colors.csv, parts.csv, and elements.csv from a Rebrickable ZIP.
    Returns dict mapping type (colors/parts/elements) to file path.
    """
    found: dict[str, str] = {}
    targets = {"colors.csv": "colors", "parts.csv": "parts", "elements.csv": "elements"}

    with zipfile.ZipFile(zip_path) as z:
        for zinfo in z.infolist():
            base = os.path.basename(zinfo.filename).lower()
            if base in targets:
                out = os.path.join(dest_dir, base)
                with z.open(zinfo) as src, open(out, "wb") as dst:
                    shutil.copyfileobj(src, dst)
                found[targets[base]] = out

    return found


def read_csv_rows(path: str, offset: int, limit: int) -> Iterator[dict[str, Any]]:
    """
    Yield up to `limit` dict-rows from a CSV file starting at `offset`.
    Works for both .csv and .csv.gz files.
    """
    opener = gzip.open if path.lower().endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        # Skip offset rows
        for _ in range(offset):
            try:
                next(reader)
            except StopIteration:
                return
        # Yield up to limit rows
        for i, row in enumerate(reader):
            if i >= limit:
                break
            yield row


# -------------------------------------------------------------------------------------------------
# Data Conversion Helpers
# -------------------------------------------------------------------------------------------------

def to_int(value: Any, default: int = 0) -> int:
    """Safely convert a value to int, returning default on failure."""
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return default


def to_str(value: Any) -> str:
    """Convert any value to a clean string, handling None/floats safely."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_key(key: str) -> str:
    """Normalize CSV header keys (lowercase, underscores, trimmed)."""
    return (key or "").strip().lower().replace(" ", "_")


# -------------------------------------------------------------------------------------------------
# CSV/Excel Import Helpers
# -------------------------------------------------------------------------------------------------

def rows_from_csv_bytes(raw: bytes) -> Iterator[dict[str, Any]]:
    """
    Parse CSV bytes into dict rows.
    Handles encoding detection, delimiter sniffing, and normalization.
    """
    # Try multiple encodings
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        raise ValueError("Unsupported text encoding")

    # Normalize line endings
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    stream = io.StringIO(text)

    # Detect CSV dialect
    try:
        sample = stream.read(2048)
        dialect = csv.Sniffer().sniff(sample)
        stream.seek(0)
    except Exception:
        stream.seek(0)
        dialect = csv.excel

    # Parse rows
    reader = csv.DictReader(stream, dialect=dialect)
    for row in reader:
        if not any(row.values()):
            continue
        yield {
            normalize_key(k): (v.strip() if isinstance(v, str) else v)
            for k, v in row.items()
        }


def rows_from_xlsx_bytes(raw: bytes) -> Iterator[dict[str, Any]]:
    """Parse XLSX bytes into dict rows (first sheet only)."""
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    # Get headers
    try:
        headers = [normalize_key(str(h) if h is not None else "") for h in next(rows_iter)]
    except StopIteration:
        return

    # Parse data rows
    for row in rows_iter:
        if row is None or not any(row):
            continue

        data = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            data[header] = val if val is not None else ""

        yield data


def rows_from_xls_bytes(raw: bytes) -> Iterator[dict[str, Any]]:
    """Parse legacy XLS bytes into dict rows (first sheet only)."""
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)

    if sheet.nrows == 0:
        return

    # Get headers
    headers = [normalize_key(str(sheet.cell_value(0, c))) for c in range(sheet.ncols)]

    # Parse data rows
    for r in range(1, sheet.nrows):
        values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if not any(v not in ("", None) for v in values):
            continue

        data = {}
        for i, header in enumerate(headers):
            val = values[i] if i < len(values) else ""
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            data[header] = str(val).strip() if isinstance(val, str) else val

        yield data


def detect_excel_type(filename: str, raw: bytes) -> str | None:
    """
    Detect Excel file type from filename or content.
    Returns 'xlsx', 'xls', or None.
    """
    lower = (filename or "").lower()

    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls"):
        return "xls"

    # Check magic bytes
    if raw.startswith(b"PK\x03\x04"):  # ZIP header (XLSX)
        return "xlsx"
    if raw.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):  # OLE header (XLS)
        return "xls"

    return None


# -------------------------------------------------------------------------------------------------
# Part Lookup Helpers
# -------------------------------------------------------------------------------------------------

def is_element_id(part_id: str) -> bool:
    """
    Heuristic to detect if a string is an element ID.
    Element IDs are numeric and typically >= 6 digits.
    """
    part_id = (part_id or "").strip()
    return part_id.isdigit() and len(part_id) >= ELEMENT_ID_MIN_LENGTH


def extract_element_id_from_url(url: str) -> str:
    """Extract element ID from Rebrickable image URL."""
    if not url:
        return ""
    match = re.search(r"/elements/(\d+)\.jpg", str(url))
    return match.group(1) if match else ""


def single_part_token(raw: str) -> str:
    """
    Return a single alphanumeric token if valid, otherwise empty string.
    Used to filter out multi-part IDs or invalid input.
    """
    token = to_str(raw)
    match = re.fullmatch(r'[0-9A-Za-z]+', token)
    return match.group(0) if match else ""


def digits_if_suffix(token: str) -> str:
    """
    If token ends with letters (e.g., '3684c'), return just the digits ('3684').
    Otherwise return the token unchanged.
    """
    match = re.fullmatch(r'(\d+)[A-Za-z]+', token)
    return match.group(1) if match else token


def fetch_rebrickable_part(token: str, api_key: str) -> tuple[dict | None, int]:
    """
    Fetch part info from Rebrickable API.
    Tries /lego/parts/{token}/, then /lego/parts/{digits}/ if token has suffix.
    Returns (json_data or None, api_call_count).
    """
    headers = {"Authorization": f"key {api_key}"}
    attempts = 0

    # Try exact token
    try:
        response = requests.get(
            f"https://rebrickable.com/api/v3/lego/parts/{token}/",
            headers=headers,
            timeout=EXTERNAL_API_TIMEOUT
        )
        attempts += 1
        if response.status_code == 200:
            return response.json(), attempts
    except requests.RequestException:
        pass

    # Try digit-only variant if token has letter suffix
    digits = digits_if_suffix(token)
    if digits != token:
        try:
            response = requests.get(
                f"https://rebrickable.com/api/v3/lego/parts/{digits}/",
                headers=headers,
                timeout=EXTERNAL_API_TIMEOUT
            )
            attempts += 1
            if response.status_code == 200:
                return response.json(), attempts
        except requests.RequestException:
            pass

    return None, attempts


def fetch_bricklink_name(token: str) -> str | None:
    """
    Scrape part name from BrickLink catalog page.
    Returns sanitized name or None if not found.
    """
    from html import unescape
    import urllib.parse

    url = f"https://www.bricklink.com/v2/catalog/catalogitem.page?P={urllib.parse.quote(token)}"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; BlockShelf/1.0)"}

    try:
        response = requests.get(url, headers=headers, timeout=EXTERNAL_API_TIMEOUT)
        if response.status_code != 200:
            return None

        html = response.text

        # Try specific item-name element
        match = re.search(r'id=["\']item-name["\'][^>]*>(.*?)</', html, re.I | re.S)
        if match:
            name = re.sub(r'\s+', ' ', unescape(match.group(1))).strip()
            if name:
                return sanitize_text(name, allow_basic_formatting=False)

        # Try og:title meta tag
        match = re.search(r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)["\']', html, re.I)
        if match:
            title = re.sub(r'\s+', ' ', unescape(match.group(1))).strip()
            if ' : ' in title:
                title = title.split(' : ', 1)[1]
            title = re.sub(r'\s*-\s*BrickLink.*$', '', title, flags=re.I).strip()
            if title:
                return sanitize_text(title, allow_basic_formatting=False)

        # Fallback to <title>
        match = re.search(r'<title>(.*?)</title>', html, re.I | re.S)
        if match:
            title = re.sub(r'\s+', ' ', unescape(match.group(1))).strip()
            if ' : ' in title:
                title = title.split(' : ', 1)[1]
            title = re.sub(r'\s*-\s*BrickLink.*$', '', title, flags=re.I).strip()
            if title:
                return sanitize_text(title, allow_basic_formatting=False)

    except requests.RequestException:
        pass

    return None


# -------------------------------------------------------------------------------------------------
# Permission & Ownership Helpers
# -------------------------------------------------------------------------------------------------

def get_owner_from_request(request: HttpRequest) -> User:
    """
    Get the inventory owner from request.
    If ?owner=N is provided and user has permission, return that user.
    Otherwise return the logged-in user.
    """
    owner = request.user
    owner_id = request.GET.get("owner")

    if not owner_id:
        return owner

    try:
        candidate = User.objects.get(pk=owner_id)
    except User.DoesNotExist:
        raise Http404("Owner not found")

    # User can view their own inventory or inventories they're a collaborator on
    if request.user.is_authenticated and candidate == request.user:
        return candidate

    # Check collaboration permission
    is_collaborator = InventoryCollab.objects.filter(
        owner=candidate,
        collaborator=request.user,
        is_active=True,
        accepted_at__isnull=False,
    ).exists()

    if not is_collaborator:
        raise Http404("Not allowed to view this inventory")

    return candidate


def get_collab_record(viewer: User, owner: User) -> InventoryCollab | None:
    """Get the collaboration record between viewer and owner, if exists."""
    if not viewer.is_authenticated or viewer == owner:
        return None

    return InventoryCollab.objects.filter(
        owner=owner,
        collaborator=viewer,
        is_active=True,
        accepted_at__isnull=False
    ).first()


def can_edit(viewer: User, owner: User) -> bool:
    """Check if viewer can edit owner's inventory."""
    if viewer == owner:
        return True

    collab = get_collab_record(viewer, owner)
    return bool(collab and collab.can_edit)


def can_delete(viewer: User, owner: User) -> bool:
    """Check if viewer can delete items in owner's inventory."""
    if viewer == owner:
        return True

    collab = get_collab_record(viewer, owner)
    return bool(collab and collab.can_delete)


# -------------------------------------------------------------------------------------------------
# IP Address Helper
# -------------------------------------------------------------------------------------------------

def get_client_ip(request: HttpRequest) -> str:
    """
    Get client IP address from request.
    Handles reverse proxy headers (X-Forwarded-For, X-Real-IP).
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        # Take the first IP in the chain (client IP)
        return x_forwarded_for.split(',')[0].strip()

    return request.META.get('REMOTE_ADDR', '0.0.0.0')
