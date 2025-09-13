# --- Standard library
import csv
import io
import os
import gzip
import zipfile
import shutil
import tempfile
import time
import re
import urllib.parse
from io import BytesIO
from html import unescape

# --- Third-party
import requests

# --- Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm
from django.core.cache import cache
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, ExpressionWrapper, IntegerField, Q
from django.db.utils import OperationalError, ProgrammingError
from django.http import (
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST

# --- Local (project)
from .models import (
    InventoryItem,
    InventoryShare,
    InventoryCollab,
    RBPart,
    RBColor,
    RBElement,
    AppConfig,
)

from .forms import (
    ProfileForm,
    InventoryItemForm,
    ImportCSVForm,
    InviteCollaboratorForm,
    InventoryImportForm,
    AppConfigForm,
)

from .utils import get_effective_config

# ---------- Inventory pages ----------

@login_required
def inventory_list(request):
    owner = _get_owner_from_request(request)
    q = request.GET.get("q", "").strip()

    # sort mapping
    sort = request.GET.get("sort", "name")
    direction = request.GET.get("dir", "asc")
    mapping = {
        "name": "name",
        "part": "part_id",
        "color": "color",
        "total": "quantity_total",
        "used": "quantity_used",
        "avail": "quantity_available",
        "loc": "storage_location",
    }
    order = mapping.get(sort, "name")
    if direction == "desc":
        order = f"-{order}"

    qs = InventoryItem.objects.filter(user=owner)
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(part_id__icontains=q)
            | Q(color__icontains=q)
            | Q(storage_location__icontains=q)
        )
    qs = qs.order_by(order, "id")

    cfg = get_effective_config()
    per_page = getattr(cfg, "items_per_page", 25) or 25
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))

    # permissions
    can_edit = _can_edit(request.user, owner)
    can_delete = _can_delete(request.user, owner)

    # inventories this user can access (for the dropdown)
    collab_list = []
    if request.user.is_authenticated:
        collab_list = list(
            InventoryCollab.objects
            .filter(collaborator=request.user, is_active=True, accepted_at__isnull=False)
            .select_related("owner")
            .order_by("owner__username")
            .values_list("owner", flat=True)
        )
        collab_list = list(User.objects.filter(id__in=collab_list).order_by("username"))

    return render(request, "inventory/inventory_list.html", {
        "page_obj": page_obj,
        "q": q,
        "current_sort": sort,
        "current_dir": direction,
        "owner_context": owner,
        "is_own_inventory": request.user == owner,
        "can_edit_inventory": can_edit,
        "can_delete_inventory": can_delete,
        "collab_list": collab_list,
        "import_form": ImportCSVForm(),
    })

@login_required
def add_item(request):
    owner = _get_owner_from_request(request)
    if not _can_edit(request.user, owner):
        return HttpResponseForbidden("You do not have permission to add items here.")

    if request.method == "POST":
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.user = owner  # IMPORTANT: save into the owner's inventory
            item.save()
            messages.success(request, "Item added.")
            # keep context (owner) when navigating back
            return redirect(f"{reverse('inventory:list')}?owner={owner.pk}")
    else:
        form = InventoryItemForm()

    return render(request, "inventory/item_form.html", {
        "form": form,
        "owner_context": owner,
    })

@login_required
def item_create(request):
    owner = _get_owner_from_request(request)
    if not _can_edit(request.user, owner):
        return HttpResponseForbidden("You do not have permission to add items here.")

    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.user = owner  # save into the OWNER’s inventory
            item.save()
            messages.success(request, 'Item added.')
            return redirect(f"{reverse('inventory:list')}?owner={owner.pk}")
    else:
        form = InventoryItemForm()

    return render(request, 'inventory/item_form.html', {'form': form, 'owner_context': owner})

@login_required
def item_update(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    owner = item.user
    if not _can_edit(request.user, owner):
        return HttpResponseForbidden("You do not have permission to edit items here.")

    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item updated.')
            return redirect(f"{reverse('inventory:list')}?owner={owner.pk}")
    else:
        form = InventoryItemForm(instance=item)
    return render(request, 'inventory/item_form.html', {'form': form, 'item': item, 'owner_context': owner})

@login_required
def item_delete(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    owner = item.user

    if not _can_delete(request.user, owner):
        messages.error(request, "You don’t have permission to delete items in this inventory.")
        return redirect(f"{reverse('inventory:list')}?owner={owner.pk}")

    if request.method == "POST":
        item.delete()
        messages.success(request, "Item deleted.")
        return redirect(f"{reverse('inventory:list')}?owner={owner.pk}")

    # Optional: if someone GETs this URL, just bounce them back to the list
    return redirect(f"{reverse('inventory:list')}?owner={owner.pk}")

def _tmpdir():
    return tempfile.mkdtemp(prefix="rebrickable_")

def _count_csv_rows(path: str) -> int:
    """
    Count data rows (excluding header). Supports .csv and .csv.gz
    """
    opener = gzip.open if path.lower().endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        try:
            next(reader)  # header
        except StopIteration:
            return 0
        return sum(1 for _ in reader)

def _extract_reb_csvs_from_zip(zip_path: str, dest_dir: str) -> dict:
    """
    From a Rebrickable ZIP, extract colors.csv, parts.csv, elements.csv (case-insensitive).
    Returns {"colors": path?, "parts": path?, "elements": path?} with only found files.
    """
    found = {}
    targets = {"colors.csv": "colors", "parts.csv": "parts", "elements.csv": "elements"}
    with zipfile.ZipFile(zip_path) as z:
        for zinfo in z.infolist():
            base = os.path.basename(zinfo.filename).lower()
            if base in targets:
                out = os.path.join(dest_dir, base)
                with z.open(zinfo) as src, open(out, "wb") as dst:
                    shutil.copyfileobj(src, dst)
                found[targets[base]] = out
    return found

def _read_csv_rows(path: str, offset: int, limit: int):
    """
    Yield up to `limit` dict-rows starting at `offset` (0-based, excluding header).
    Works for .csv and .csv.gz. Header is detected from the file itself.
    """
    opener = gzip.open if path.lower().endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        # skip 'offset' rows
        for _ in range(offset):
            try:
                next(reader)
            except StopIteration:
                return
        # yield up to 'limit'
        for i, row in enumerate(reader):
            if i >= limit:
                break
            yield row

@login_required
@require_POST
def reb_bootstrap_prepare(request):
    """
    Admin-only. Accept an upload (ZIP, CSV, CSV.GZ), stash temp files, count rows,
    remember job metadata in cache, and return counts + job id.
    """
    if not request.user.is_staff:
        return HttpResponseForbidden("Admins only.")

    upload = request.FILES.get("dataset_file")
    if not upload:
        return JsonResponse({"ok": False, "error": "No file uploaded."}, status=400)

    temp_root = _tmpdir()
    src_path = os.path.join(temp_root, upload.name)
    with open(src_path, "wb") as dst:
        for chunk in upload.chunks():
            dst.write(chunk)

    files = {}
    # ZIP → extract wanted CSVs
    if zipfile.is_zipfile(src_path):
        files = _extract_reb_csvs_from_zip(src_path, temp_root)
        # optional: remove original zip to save space
        try: os.remove(src_path)
        except OSError: pass
    else:
        # Single file (csv or csv.gz): guess kind from filename
        lname = os.path.basename(src_path).lower()
        if lname.endswith((".csv", ".csv.gz")):
            if "color" in lname and "part" not in lname and "element" not in lname:
                files["colors"] = src_path
            elif "element" in lname:
                files["elements"] = src_path
            else:
                # default assume 'parts'
                files["parts"] = src_path
        else:
            shutil.rmtree(temp_root, ignore_errors=True)
            return JsonResponse({"ok": False, "error": "Unsupported file type."}, status=400)

    totals = {}
    for kind, path in files.items():
        try:
            totals[kind] = _count_csv_rows(path)
        except Exception as e:
            totals[kind] = 0

    # Persist job info in cache
    job_id = get_random_string(12)
    cache.set(
        f"rebjob:{job_id}",
        {
            "root": temp_root,
            "files": files,   # kind -> path
            "totals": totals, # kind -> int
        },
        60 * 60,  # 1 hour
    )

    return JsonResponse({"ok": True, "job": job_id, "totals": totals})


@login_required
@require_POST
def reb_bootstrap_run(request):
    """
    Admin-only. Import a batch for a given job/kind at the given offset.
    POST: job, kind (colors|parts|elements), offset, batch_size
    """
    if not request.user.is_staff:
        return HttpResponseForbidden("Admins only.")

    job = (request.POST.get("job") or "").strip()
    kind = (request.POST.get("kind") or "").strip().lower()
    try:
        offset = int(request.POST.get("offset", "0"))
    except Exception:
        offset = 0
    try:
        batch_size = int(request.POST.get("batch_size", "2000"))
    except Exception:
        batch_size = 2000
    batch_size = max(1, min(10000, batch_size))

    jobkey = f"rebjob:{job}"
    meta = cache.get(jobkey)
    if not meta:
        return JsonResponse({"ok": False, "error": "Job not found or expired."}, status=400)

    files = meta.get("files", {})
    totals = meta.get("totals", {})
    path = files.get(kind)
    if not path:
        return JsonResponse({"ok": False, "error": f"No file for kind '{kind}'."}, status=400)

    total = int(totals.get(kind, 0))

    processed = 0
    created = 0
    updated = 0
    skipped = 0
    messages_log = []

    # Upsert helpers
    def upsert_color(row):
        nonlocal created, updated, skipped
        try:
            cid = int((row.get("id") or row.get("color_id") or "").strip())
        except Exception:
            skipped += 1
            return
        name = (row.get("name") or "").strip()
        obj, was_created = RBColor.objects.update_or_create(
            id=cid, defaults={"name": name or ""}
        )
        if was_created:
            created += 1
        else:
            updated += 1

    def upsert_part(row):
        nonlocal created, updated, skipped
        part_num = (row.get("part_num") or "").strip()
        if not part_num:
            skipped += 1
            return
        name = (row.get("name") or "").strip()
        obj, was_created = RBPart.objects.update_or_create(
            part_num=part_num,
            defaults={"name": name or obj.name if 'obj' in locals() else name or ""},
        )
        if was_created:
            created += 1
        else:
            # if name empty in CSV don’t clobber non-empty DB name
            updated += 1

    def upsert_element(row):
        nonlocal created, updated, skipped
        element_id = (row.get("element_id") or "").strip()
        part_num   = (row.get("part_num") or "").strip()
        color_id   = (row.get("color_id") or "").strip()

        if not (element_id and part_num and color_id):
            skipped += 1
            return
        try:
            cid = int(color_id)
        except Exception:
            skipped += 1
            return

        # Ensure RBPart and RBColor exist
        part, _ = RBPart.objects.get_or_create(part_num=part_num, defaults={"name": ""})
        color, _ = RBColor.objects.get_or_create(id=cid, defaults={"name": ""})

        obj, was_created = RBElement.objects.update_or_create(
            element_id=element_id, defaults={"part": part, "color": color}
        )
        if was_created:
            created += 1
        else:
            updated += 1

    # Process a slice
    try:
        for row in _read_csv_rows(path, offset, batch_size):
            if kind == "colors":
                upsert_color(row)
            elif kind == "parts":
                upsert_part(row)
            elif kind == "elements":
                upsert_element(row)
            else:
                skipped += 1
            processed += 1
    except Exception as e:
        messages_log.append(f"Parse error: {e}")

    new_offset = offset + processed
    done = new_offset >= total

    # If this kind is done and *all* kinds are present & finished, cleanup temp dir + cache.
    if done:
        # Check if all kinds done by comparing requests from UI; we don't track per-kind offsets
        # so we keep the files until the UI tells us to delete, or rely on TTL.
        pass

    return JsonResponse({
        "ok": True,
        "done": done,
        "kind": kind,
        "offset": new_offset,
        "total": total,
        "processed": processed,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "messages": messages_log[:200],
    })


# ---------- Import / Export ----------

@login_required
def export_csv(request):
    """Export current user's inventory to CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(['name', 'part_id', 'color', 'quantity_total', 'quantity_used',
                     'storage_location', 'image_url', 'notes'])
    for item in InventoryItem.objects.filter(user=request.user).order_by('name'):
        writer.writerow([
            item.name,
            item.part_id,
            item.color,
            item.quantity_total,
            item.quantity_used,
            item.storage_location,
            item.image_url,
            (item.notes or '').replace('\n', ' ').strip(),
        ])
    return response


def _to_int(v, default=0):
    try:
        return int(str(v).strip())
    except Exception:
        return default

def _to_str(v) -> str:
    """Return a clean string for text fields, handling ints/floats/None safely."""
    if v is None:
        return ""
    # Keep integers clean (avoid '3001.0')
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()

def _norm_key(k: str) -> str:
    return (k or "").strip().lower().replace(" ", "_")

def _rows_from_csv_bytes(raw: bytes):
    """Yield dict rows from CSV bytes, with delimiter sniffing and encoding fallbacks."""
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Unsupported text encoding")

    text = text.replace("\r\n", "\n").replace("\r", "\n")
    stream = io.StringIO(text)

    try:
        sample = stream.read(2048)
        dialect = csv.Sniffer().sniff(sample)
        stream.seek(0)
    except Exception:
        stream.seek(0)
        dialect = csv.excel

    reader = csv.DictReader(stream, dialect=dialect)
    for row in reader:
        if not any(row.values()):
            continue
        yield {_norm_key(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}

def _rows_from_xlsx_bytes(raw: bytes):
    """Yield dict rows from XLSX bytes using openpyxl (first sheet)."""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [_norm_key(str(h) if h is not None else "") for h in next(rows_iter)]
    except StopIteration:
        return
    for row in rows_iter:
        if row is None or not any(row):
            continue
        d = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            # Normalize numbers so 3001.0 becomes 3001
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            # Leave as-is; we'll coerce to str where needed in import_csv
            d[h] = val if val is not None else ""
        yield d


def _rows_from_xls_bytes(raw: bytes):
    """Yield dict rows from legacy XLS bytes using xlrd (first sheet)."""
    import xlrd
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    if sh.nrows == 0:
        return
    headers = [ _norm_key(str(sh.cell_value(0, c))) for c in range(sh.ncols) ]
    for r in range(1, sh.nrows):
        values = [ sh.cell_value(r, c) for c in range(sh.ncols) ]
        if not any(v not in ("", None) for v in values):
            continue
        d = {}
        for i, h in enumerate(headers):
            v = values[i] if i < len(values) else ""
            # xlrd returns numbers as floats; try to keep whole numbers clean
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            d[h] = str(v).strip() if isinstance(v, str) else v
        yield d

def _detect_excel_kind(name: str, raw: bytes):
    """Return 'xlsx', 'xls', or None."""
    lower = (name or "").lower()
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls"):
        return "xls"
    # Fallback by magic bytes:
    if raw.startswith(b"PK\x03\x04"):  # zip header, typical for .xlsx
        return "xlsx"
    # OLE Compound File magic for .xls: D0 CF 11 E0 A1 B1 1A E1
    if raw.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls"
    return None

@login_required
@require_POST
def import_csv(request):
    form = ImportCSVForm(request.POST, request.FILES)
    if not form.is_valid():
        # Re-render list with errors
        q = (request.GET.get('q') or '').strip()
        items = InventoryItem.objects.filter(user=request.user).order_by('name')
        page_obj = Paginator(items, 25).get_page(request.GET.get('page'))
        return render(request, 'inventory/inventory_list.html', {
            'page_obj': page_obj,
            'q': q,
            'import_form': form,
        })

    upload = form.cleaned_data['file']
    raw = upload.read()

    kind = _detect_excel_kind(upload.name, raw)

    # Choose a row generator
    try:
        if kind == "xlsx":
            rows = _rows_from_xlsx_bytes(raw)
        elif kind == "xls":
            rows = _rows_from_xls_bytes(raw)
        else:
            rows = _rows_from_csv_bytes(raw)
    except Exception as e:
        messages.error(request, f"Could not read file: {e}")
        return redirect('inventory:list')

    added = updated = skipped = 0
    dupe_keys = 0

    with transaction.atomic():
        for row in rows:
            # Normalize keys
            data = { _norm_key(k): v for k, v in row.items() }

            # Skip blank
            if not any(v not in ("", None) for v in data.values()):
                continue

            name  = _to_str(data.get('name'))
            part  = _to_str(data.get('part_id'))
            color = _to_str(data.get('color'))
            qtot  = _to_int(data.get('quantity_total'), 0)
            quse  = _to_int(data.get('quantity_used'), 0)
            loc   = _to_str(data.get('storage_location'))
            img   = _to_str(data.get('image_url'))
            notes = _to_str(data.get('notes'))


            # Need at least a part or a name
            if not part and not name:
                skipped += 1
                continue

            qs = InventoryItem.objects.filter(
                user=request.user,
                part_id=part,
                color=color,
            )

            if not qs.exists():
                InventoryItem.objects.create(
                    user=request.user,
                    part_id=part,
                    color=color,
                    name=name or 'Unknown',
                    quantity_total=qtot,
                    quantity_used=quse,
                    storage_location=loc,
                    image_url=img,
                    notes=notes,
                )
                added += 1
            else:
                if qs.count() > 1:
                    dupe_keys += 1
                obj = qs.order_by('id').first()
                if name:
                    obj.name = name
                if data.get('quantity_total') not in (None, ''):
                    obj.quantity_total = qtot
                if data.get('quantity_used') not in (None, ''):
                    obj.quantity_used = quse
                if loc:
                    obj.storage_location = loc
                if img:
                    obj.image_url = img
                if notes:
                    if obj.notes and notes not in obj.notes:
                        obj.notes = (obj.notes + "\n" + notes).strip()
                    else:
                        obj.notes = notes or obj.notes
                obj.save()
                updated += 1

    msg = f"Import complete: {added} added, {updated} updated, {skipped} skipped."
    if dupe_keys:
        msg += f" (Note: {dupe_keys} duplicate key set(s) existed; updated the first match.)"
    messages.success(request, msg)
    return redirect('inventory:list')


# ---------- Local-first part lookup (with API fallback, caching, and throttling) ----------


LOOKUP_LIMIT_PER_MIN = 60         # per-user
LOOKUP_WINDOW_SECONDS = 60
LOOKUP_CACHE_TTL = 60 * 60 * 24   # 1 day

def _rate_limit_ok(user_id: int) -> bool:
    """Simple sliding window counter stored in cache."""
    bucket = int(time.time() // LOOKUP_WINDOW_SECONDS)
    key = f"rl:lookup:{user_id}:{bucket}"
    cache.add(key, 0, timeout=LOOKUP_WINDOW_SECONDS)  # create if missing
    try:
        current = cache.incr(key)
    except ValueError:
        cache.set(key, 1, LOOKUP_WINDOW_SECONDS)
        current = 1
    return current <= LOOKUP_LIMIT_PER_MIN

def _is_element_id(s: str) -> bool:
    """Heuristic: BLOCKSHELF element IDs are numeric and typically >= 6 digits."""
    s = (s or "").strip()
    return s.isdigit() and len(s) >= 6

def _extract_element_id_from_url(url: str) -> str:
    import re
    if not url:
        return ""
    m = re.search(r"/elements/(\d+)\.jpg", str(url))
    return m.group(1) if m else ""

@login_required
def lookup_part(request):
    """
    Lookup by part number OR element ID.
    Order: cache → local DB → (throttle) → Rebrickable API → persist → cache.
    Returns: { ok, found, resolved: 'part'|'element', name, part_id, image_url, [color], [element_id] }

    Enhancement: If Rebrickable can't resolve a PART id (or we're rate-limited),
    we fall back to BrickLink's public catalog page to grab the NAME only.
    """
    raw = (request.GET.get('part_id') or '').strip()
    if not raw:
        return JsonResponse({'ok': False, 'error': 'Missing part_id'}, status=400)

    cfg = get_effective_config()
    api_key = cfg.rebrickable_api_key or getattr(settings, "REBRICKABLE_API_KEY", "")

    is_element = _is_element_id(raw)
    cache_key = f"lookup:{'element' if is_element else 'part'}:{raw}"
    cached = cache.get(cache_key)
    if cached:
        return JsonResponse(cached)

    # -------------------- ELEMENT PATH --------------------
    if is_element:
        # 1) Local RBElement (preferred)
        el = RBElement.objects.select_related('part', 'color').filter(element_id=str(raw)).first()
        if el:
            rbpart = el.part
            img = (rbpart.image_url or '').strip()

            # Try to backfill part image once, if missing
            if not img and _rate_limit_ok(request.user.id) and api_key:
                try:
                    r = requests.get(
                        f"https://rebrickable.com/api/v3/lego/parts/{rbpart.part_num}/",
                        headers={'Authorization': f'key {api_key}'},
                        timeout=10
                    )
                    if r.ok:
                        img = (r.json().get('part_img_url') or '').strip()
                        if img:
                            RBPart.objects.filter(part_num=rbpart.part_num).update(image_url=img)
                except requests.RequestException:
                    pass

            data = {
                'ok': True, 'found': True, 'resolved': 'element',
                'name': rbpart.name,
                'part_id': rbpart.part_num,
                'color': el.color.name,
                'image_url': img,
                'element_id': str(raw),
            }
            cache.set(cache_key, data, LOOKUP_CACHE_TTL)
            return JsonResponse(data)

        # 2) External API fallback: /elements/{id}/
        try:
            r = requests.get(
                f"https://rebrickable.com/api/v3/blockshelf/elements/{raw}/",
                headers={'Authorization': f'key {api_key}'}, timeout=10
            )
            if r.status_code == 404:
                miss = {'ok': True, 'found': False}
                cache.set(cache_key, miss, 300)
                return JsonResponse(miss)
            r.raise_for_status()
            j = r.json()

            part = j.get('part') or {}
            color = j.get('color') or {}
            part_num = (part.get('part_num') or '').strip()
            name = (part.get('name') or '').strip()
            part_img = (part.get('part_img_url') or '').strip()
            color_id = color.get('id')
            color_name = (color.get('name') or '').strip()

            rbpart = None
            if part_num:
                rbpart, _ = RBPart.objects.update_or_create(
                    part_num=part_num,
                    defaults={'name': name, 'image_url': part_img}
                )
            rbcolor = None
            if color_id is not None:
                try:
                    cid = int(color_id)
                    rbcolor, _ = RBColor.objects.update_or_create(
                        id=cid, defaults={'name': color_name}
                    )
                except Exception:
                    rbcolor = None

            if rbpart and rbcolor:
                RBElement.objects.update_or_create(
                    element_id=str(raw), defaults={'part': rbpart, 'color': rbcolor}
                )

            data = {
                'ok': True, 'found': bool(part_num), 'resolved': 'element',
                'name': name,
                'part_id': part_num or str(raw),
                'color': color_name,
                'image_url': part_img,
                'element_id': str(raw),
            }
            cache.set(cache_key, data, LOOKUP_CACHE_TTL)
            return JsonResponse(data)

        except requests.RequestException as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=502)

    # -------------------- PART PATH (single token; trailing-letter fallback + BrickLink name fallback) --------------------
    token = _single_part_token(raw)
    if not token:
        # e.g. "30293 & 30294" or "N/A" → not a single token → treat as not found
        miss = {'ok': True, 'found': False}
        cache.set(cache_key, miss, 300)
        return JsonResponse(miss)

    # 1) Local RBPart: try token, then digits-only if token ends with letters
    name = ''
    img = ''
    pn = token

    rb = RBPart.objects.filter(part_num=token).first()
    if rb:
        pn = rb.part_num
        name = (rb.name or '')
        img = (rb.image_url or '')

    if not rb:
        digits = _digits_if_suffix(token)
        if digits != token:
            rb2 = RBPart.objects.filter(part_num=digits).first()
            if rb2:
                pn = rb2.part_num
                name = name or (rb2.name or '')
                img = img or (rb2.image_url or '')

    # 2) Rebrickable API fallback (tries token, then digits-only)
    if (not name or not img):
        if _rate_limit_ok(request.user.id) and api_key:
            try:
                res, _attempts = _rb_fetch_part_simple(token, api_key)
                if res:
                    pn = res.get('part_num') or pn
                    name = name or (res.get('name') or '')
                    img = img or (res.get('part_img_url') or '')
                    RBPart.objects.update_or_create(
                        part_num=pn, defaults={'name': name, 'image_url': img}
                    )
            except requests.RequestException:
                pass

    # 3) BrickLink fallback — NAME ONLY (no image)
    if not name:
        bl_name = _bricklink_name_for_part(token)
        if not bl_name:
            # Try digits-only if token had letter suffix (e.g. 3684c → 3684)
            digits = _digits_if_suffix(token)
            if digits != token:
                bl_name = _bricklink_name_for_part(digits)
        if bl_name:
            name = bl_name  # do NOT set image; BL doesn't give one here

    # Final response for part path
    if not (name or img):
        # Nothing found on either RB or BL
        miss = {'ok': True, 'found': False}
        cache.set(cache_key, miss, 300)
        return JsonResponse(miss)

    data = {
        'ok': True, 'found': True, 'resolved': 'part',
        'name': name,
        'part_id': pn,
        'image_url': img,
        'element_id': _extract_element_id_from_url(img),
    }
    cache.set(cache_key, data, LOOKUP_CACHE_TTL)
    return JsonResponse(data)



@login_required
def settings_page(request):
    """Render the settings / maintenance page."""
    return render(request, "inventory/settings.html", {})


@login_required
@require_GET
def bulk_update_missing_preview(request):
    """
    Count items that need updating AND have exactly one alphanumeric part_id.
    Also report how many will be skipped due to multi-IDs and a few examples.
    """
    try:
        base = (InventoryItem.objects
                .filter(user=request.user)
                .filter(Q(name__iexact="unknown") | Q(image_url__isnull=True) | Q(image_url__exact=""))
                .exclude(part_id__isnull=True).exclude(part_id__exact=""))

        eligible = 0
        skipped_multi = 0
        examples = []
        for it in base.only("id", "part_id"):
            token = _single_part_token(it.part_id)
            if token:
                eligible += 1
            else:
                skipped_multi += 1
                if len(examples) < 6:
                    examples.append({"id": it.id, "part_id": _to_str(it.part_id)})

        return JsonResponse({
            "count": eligible,
            "skipped_multi": skipped_multi,
            "skipped_examples": examples,
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
@require_POST
def bulk_update_missing_batch(request):
    try:
        after_id = int(request.POST.get("after_id", 0))
    except Exception:
        after_id = 0

    try:
        batch_size = int(request.POST.get("batch_size", 50))
    except Exception:
        batch_size = 50
    batch_size = max(1, min(500, batch_size))

    qs = (InventoryItem.objects
          .filter(user=request.user)
          .filter(Q(name__iexact="unknown") | Q(image_url__isnull=True) | Q(image_url__exact=""))
          .exclude(part_id__isnull=True).exclude(part_id__exact="")
          .filter(id__gt=after_id)
          .order_by("id")[:batch_size])

    items = list(qs)
    if not items:
        return JsonResponse({
            "done": True, "last_id": after_id, "processed": 0,
            "updated_names": 0, "updated_images": 0, "skipped": 0, "api_calls": 0,
            "messages": [],
        })

    cfg = get_effective_config()
    api_key = cfg.rebrickable_api_key or getattr(settings, "REBRICKABLE_API_KEY", "")

    updated_names = 0
    updated_images = 0
    skipped = 0
    api_calls = 0
    last_id = after_id
    messages_log = []

    # NEW: count only eligible (single-token) items so the progress matches Scan
    processed_eligible = 0

    for item in items:
        last_id = item.id

        need_name = (item.name or "").strip().lower() == "unknown"
        need_img  = not (item.image_url or "").strip()
        if not (need_name or need_img):
            skipped += 1
            continue

        token = _single_part_token(item.part_id)
        if not token:
            skipped += 1
            messages_log.append(f"Skipped (multiple IDs): #{item.id} '{_to_str(item.part_id)}'")
            continue

        # count only when we actually process an eligible row
        processed_eligible += 1

        part_name = ""
        part_img  = ""

        # Local RBPart (token, then digits fallback)
        rb = RBPart.objects.filter(part_num=token).first()
        if rb:
            part_name = (rb.name or "").strip()
            part_img  = (rb.image_url or "").strip()
        if not rb:
            digits = _digits_if_suffix(token)
            if digits != token:
                rb2 = RBPart.objects.filter(part_num=digits).first()
                if rb2:
                    part_name = part_name or (rb2.name or "").strip()
                    part_img  = part_img  or (rb2.image_url or "").strip()

        # Rebrickable API if still needed
        used_bricklink = False
        if ((need_name and not part_name) or (need_img and not part_img)) and api_key and _rate_limit_ok(request.user.id):
            try:
                res, attempts = _rb_fetch_part_simple(token, api_key)
                api_calls += attempts
                if res:
                    pn = res.get("part_num") or token
                    nm = (res.get("name") or "").strip()
                    img = (res.get("part_img_url") or "").strip()
                    part_name = part_name or nm
                    part_img  = part_img  or img
                    RBPart.objects.update_or_create(
                        part_num=pn, defaults={"name": nm, "image_url": img}
                    )
            except requests.RequestException:
                pass

        # BrickLink fallback for NAME ONLY (no image) if RB failed to provide name
        if need_name and not part_name:
            bl_name = _bricklink_name_for_part(token)
            if not bl_name:
                # try digits fallback if suffix
                digits = _digits_if_suffix(token)
                if digits != token:
                    bl_name = _bricklink_name_for_part(digits)
            if bl_name:
                part_name = bl_name
                used_bricklink = True
                messages_log.append(
                    f"BrickLink fallback: #{item.id} part '{token}' → name='{bl_name}'. No image/buy available."
                )
            else:
                messages_log.append(
                    f"Not found on Rebrickable and BrickLink: #{item.id} part '{token}'."
                )

        changed = False
        if need_name and part_name:
            item.name = part_name
            updated_names += 1
            changed = True
            # small, helpful line in the log
            if used_bricklink:
                messages_log.append(f"Updated name via BrickLink: #{item.id} '{token}' → '{part_name}'")
            else:
                messages_log.append(f"Updated name: #{item.id} '{token}' → '{part_name}'")

        # Only fill image when empty; BrickLink never supplies images here
        if need_img and part_img:
            item.image_url = part_img
            updated_images += 1
            changed = True
            messages_log.append(f"Updated image from Rebrickable: #{item.id} '{token}'")
            try:
                m = re.search(r"/elements/(\d+)\.jpg", part_img or "")
                if hasattr(item, "element_id") and not (item.element_id or "") and m:
                    item.element_id = m.group(1)
            except Exception:
                pass

        if changed:
            item.save()
        else:
            skipped += 1

    return JsonResponse({
        "done": False,
        "last_id": last_id,
        # NEW: report only eligible processed count
        "processed": processed_eligible,
        "updated_names": updated_names,
        "updated_images": updated_images,
        "skipped": skipped,
        "api_calls": api_calls,
        "messages": messages_log[:200],  # keep response small
    })


def _single_part_token(raw: str) -> str:
    """Return ONE contiguous [0-9A-Za-z]+ token; otherwise '' (multi IDs or junk)."""
    s = _to_str(raw)
    m = re.fullmatch(r'[0-9A-Za-z]+', s)
    return m.group(0) if m else ""

def _digits_if_suffix(token: str) -> str:
    """If token ends with letters (e.g. 3684c) return digits (3684), else token."""
    m = re.fullmatch(r'(\d+)[A-Za-z]+', token)
    return m.group(1) if m else token

def _rb_fetch_part_simple(token: str, api_key: str):
    """
    Try /blockshelf/parts/{token}/, then /blockshelf/parts/{digits}/ if token has a letter suffix.
    Returns (json_or_None, api_attempts).
    """
    headers = {"Authorization": f"key {api_key}"}
    attempts = 0
    try:
        r = requests.get(f"https://rebrickable.com/api/v3/lego/parts/{token}/", headers=headers, timeout=10)
        attempts += 1
        if r.status_code == 200:
            return r.json(), attempts
    except requests.RequestException:
        pass

    digits = _digits_if_suffix(token)
    if digits != token:
        try:
            r = requests.get(f"https://rebrickable.com/api/v3/lego/parts/{digits}/", headers=headers, timeout=10)
            attempts += 1
            if r.status_code == 200:
                return r.json(), attempts
        except requests.RequestException:
            pass

    return None, attempts

def _bricklink_name_for_part(token: str) -> str | None:
    """
    Best-effort scrape of the BrickLink catalog page title for a part name.
    Example page: https://www.bricklink.com/v2/catalog/catalogitem.page?P=3861
    Returns the cleaned name or None.
    """
    url = f"https://www.bricklink.com/v2/catalog/catalogitem.page?P={urllib.parse.quote(token)}"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; BrickShelf/1.0)"}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code != 200:
            return None
        html = r.text

        # Try specific label (if present)
        m = re.search(r'id=["\']item-name["\'][^>]*>(.*?)</', html, re.I | re.S)
        if m:
            name = re.sub(r'\s+', ' ', unescape(m.group(1))).strip()
            if name:
                return name

        # Try og:title
        m = re.search(r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)["\']', html, re.I)
        if m:
            title = re.sub(r'\s+', ' ', unescape(m.group(1))).strip()
            if ' : ' in title:
                title = title.split(' : ', 1)[1]
            title = re.sub(r'\s*-\s*BrickLink.*$', '', title, flags=re.I).strip()
            if title:
                return title

        # Fallback to <title>
        m = re.search(r'<title>(.*?)</title>', html, re.I | re.S)
        if m:
            title = re.sub(r'\s+', ' ', unescape(m.group(1))).strip()
            if ' : ' in title:
                title = title.split(' : ', 1)[1]
            title = re.sub(r'\s*-\s*BrickLink.*$', '', title, flags=re.I).strip()
            if title:
                return title
    except requests.RequestException:
        pass
    return None

def _extract_element_id_from_url(url: str):
    m = re.search(r"/elements/(\d+)\.jpg", url or "")
    return m.group(1) if m else None


@login_required
def settings_view(request):
    tab = request.GET.get("tab", "maintenance")

    # Always useful context
    ctx = {
        "active_tab": tab,
        "item_count": InventoryItem.objects.filter(user=request.user).count(),
    }

    # ---------- Account tab ----------
    if tab == "account":
        if request.method == "POST":
            action = request.POST.get("action")

            if action in ("save_profile", "update_profile"):
                profile_form = ProfileForm(request.POST, instance=request.user)
                password_form = PasswordChangeForm(user=request.user)  # blank
                if profile_form.is_valid():
                    profile_form.save()
                    messages.success(request, "Profile updated.")
                    return redirect(f"{reverse('inventory:settings')}?tab=account")
                else:
                    messages.error(request, "Please correct the errors below.")
                ctx.update({"profile_form": profile_form, "password_form": password_form})
                return render(request, "inventory/settings.html", ctx)

            elif action == "change_password":
                profile_form = ProfileForm(instance=request.user)  # blank
                password_form = PasswordChangeForm(user=request.user, data=request.POST)
                if password_form.is_valid():
                    user = password_form.save()
                    update_session_auth_hash(request, user)  # keep session valid
                    messages.success(request, "Password changed.")
                    return redirect(f"{reverse('inventory:settings')}?tab=account")
                else:
                    messages.error(request, "Please correct the errors below.")
                ctx.update({"profile_form": profile_form, "password_form": password_form})
                return render(request, "inventory/settings.html", ctx)

            else:
                # Unknown action; just go back to Account tab
                return redirect(f"{reverse('inventory:settings')}?tab=account")

        # GET – prefill forms
        ctx.update({
            "profile_form": ProfileForm(instance=request.user),
            "password_form": PasswordChangeForm(user=request.user),
        })
        return render(request, "inventory/settings.html", ctx)

    # ---------- Sharing tab ----------
    if tab == "sharing":
        share = InventoryShare.objects.filter(user=request.user, is_active=True).first()
        share_url = None
        if share:
            share_url = request.build_absolute_uri(
                reverse("inventory:shared_inventory", args=[share.token])
            )
        ctx.update({"share_url": share_url})
        return render(request, "inventory/settings.html", ctx)

    # ---------- Invite tab ----------
    if tab == "invite":
        ctx.update({
            "invite_form": InviteCollaboratorForm(),
            "invites": InventoryCollab.objects.filter(owner=request.user).order_by("-created_at"),
        })
        return render(request, "inventory/settings.html", ctx)

    # ---------- Site settings tab (admin only) ----------
    if tab == "config":
        if not request.user.is_staff:
            return HttpResponseForbidden("Admins only.")

        cfg = AppConfig.get_solo()  # singleton row

        if request.method == "POST":
            form = AppConfigForm(request.POST, instance=cfg)
            if form.is_valid():
                form.save()
                messages.success(request, "Site settings saved.")
                return redirect(f"{reverse('inventory:settings')}?tab=config")
            messages.error(request, "Please correct the errors below.")
        else:
            form = AppConfigForm(instance=cfg)
            if not cfg.rebrickable_api_key and getattr(settings, "REBRICKABLE_API_KEY", ""):
                form.fields["rebrickable_api_key"].initial = settings.REBRICKABLE_API_KEY

        # Optional: show if Google OAuth is configured
        google_ready = bool(
            getattr(settings, "SOCIALACCOUNT_PROVIDERS", {})
            .get("google", {})
            .get("APP", {})
            .get("client_id")
        )

        ctx.update({"appconfig_form": form, "google_ready": google_ready})
        return render(request, "inventory/settings.html", ctx)

    # ---------- Maintenance / Danger / default ----------
    return render(request, "inventory/settings.html", ctx)

    
@login_required
@require_POST
def create_invite(request):
    form = InviteCollaboratorForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Please correct the errors below.")
        return render(request, "inventory/settings.html", {
            "active_tab": "invite",
            "invite_form": form,
            "invites": InventoryCollab.objects.filter(owner=request.user).order_by("-created_at"),
        })

    email = form.cleaned_data["email"].strip().lower()
    can_edit = form.cleaned_data.get("can_edit", True)
    can_delete = form.cleaned_data.get("can_delete", False)

    if request.user.email and request.user.email.lower() == email:
        messages.error(request, "You cannot invite yourself.")
        return redirect(f"{reverse('inventory:settings')}?tab=invite")

    inv = InventoryCollab.objects.create(
        owner=request.user,
        invited_email=email,
        can_edit=can_edit,
        can_delete=can_delete,
    )
    # (Optional) send email here if EMAIL_* configured, otherwise UI will show the link.
    messages.success(request, "Invite created. Copy the link from the table below.")
    return redirect(f"{reverse('inventory:settings')}?tab=invite")

@login_required
def revoke_invite(request, pk):
    inv = get_object_or_404(InventoryCollab, pk=pk, owner=request.user, is_active=True)
    inv.is_active = False
    inv.save(update_fields=["is_active"])
    messages.success(request, "Invite revoked.")
    return redirect(f"{reverse('inventory:settings')}?tab=invite")

def accept_invite(request, token):
    inv = get_object_or_404(InventoryCollab, token=token, is_active=True)
    # already accepted? bounce to the owner’s inventory (after login)
    if inv.accepted_at:
        if not request.user.is_authenticated:
            return redirect(reverse("login") + f"?next={request.path}")
        return redirect(f"{reverse('inventory:list')}?owner={inv.owner_id}")

    # require login to bind the invite
    if not request.user.is_authenticated:
        return redirect(reverse("login") + f"?next={request.path}")

    if request.user == inv.owner:
        messages.error(request, "You cannot accept your own invite.")
        return redirect(f"{reverse('inventory:settings')}?tab=invite")

    # If they already have an active collab, just update perms; otherwise accept this invite.
    existing = InventoryCollab.objects.filter(
        owner=inv.owner, collaborator=request.user, is_active=True, accepted_at__isnull=False
    ).first()
    if existing:
        existing.can_edit = inv.can_edit or existing.can_edit
        existing.can_delete = inv.can_delete or existing.can_delete
        existing.save(update_fields=["can_edit", "can_delete"])
        # retire the invite
        inv.is_active = False
        inv.save(update_fields=["is_active"])
    else:
        inv.mark_accepted(request.user)

    messages.success(request, f"You now have access to {inv.owner.username}'s inventory.")
    return redirect(f"{reverse('inventory:list')}?owner={inv.owner_id}")

@login_required
@require_POST
def update_invite(request, pk):
    """Update can_edit / can_delete for an active invite/collaboration."""
    inv = get_object_or_404(InventoryCollab, pk=pk, owner=request.user)
    if not inv.is_active:
        messages.error(request, "Cannot edit a revoked invite.")
        return redirect(f"{reverse('inventory:settings')}?tab=invite")

    inv.can_edit = bool(request.POST.get("can_edit"))
    inv.can_delete = bool(request.POST.get("can_delete"))
    inv.save(update_fields=["can_edit", "can_delete"])
    messages.success(request, "Permissions updated.")
    return redirect(f"{reverse('inventory:settings')}?tab=invite")

@login_required
@require_POST
def purge_invite(request, pk):
    """Permanently delete a revoked invite from the list."""
    inv = get_object_or_404(InventoryCollab, pk=pk, owner=request.user)
    if inv.is_active:
        messages.error(request, "You can only delete revoked invites.")
        return redirect(f"{reverse('inventory:settings')}?tab=invite")

    inv.delete()
    messages.success(request, "Revoked invite removed.")
    return redirect(f"{reverse('inventory:settings')}?tab=invite")

@login_required
def inventory_switcher(request):
    my_inv = request.user
    collabs = (InventoryCollab.objects
               .filter(collaborator=request.user, is_active=True, accepted_at__isnull=False)
               .select_related("owner")
               .order_by("owner__username"))
    return render(request, "inventory/inventory_switcher.html", {
        "my_inv": my_inv,
        "collabs": collabs,
    })

@login_required
@require_POST
def delete_all_inventory(request):
    """Danger Zone: purge all of the authenticated user's items."""
    confirm_text = request.POST.get("confirm_text", "")
    confirm_ack = request.POST.get("confirm_ack") == "on"

    if confirm_text != "DELETE" or not confirm_ack:
        messages.error(request, "Confirmation failed. Type DELETE and tick the checkbox.")
        return redirect(f"{reverse('inventory:settings')}?tab=danger")

    # Delete ONLY this user's inventory.
    qs = InventoryItem.objects.filter(user=request.user)
    deleted_count = qs.count()
    qs.delete()

    messages.success(request, f"Deleted {deleted_count} item(s) from your inventory.")
    return redirect("inventory:list")


# ---------- Signup ----------

def signup(request):
    """Username/password signup controlled by AppConfig.allow_registration."""
    if not get_effective_config().allow_registration:
        messages.error(request, "Self-registration is disabled. Please contact the administrator.")
        return redirect("login")

    if request.user.is_authenticated:
        return redirect('inventory:list')

    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Multiple auth backends are installed (Django + allauth) → specify backend.
            auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, 'Welcome! Your account has been created.')
            return redirect('inventory:list')
    else:
        form = UserCreationForm()

    return render(request, 'signup.html', {'form': form})


@login_required
def check_duplicate(request):
    """Return whether a duplicate (same part_id+color for this user) exists."""
    part_id = (request.GET.get("part_id") or "").strip()
    color = (request.GET.get("color") or "").strip()
    exclude = request.GET.get("exclude")  # pk to exclude (editing case)

    qs = InventoryItem.objects.filter(
        user=request.user,
        part_id__iexact=part_id,
        color__iexact=color,
    )
    if exclude:
        qs = qs.exclude(pk=exclude)

    exists = qs.exists()
    data = {"exists": exists}
    if exists:
        data["count"] = qs.count()
        data["items"] = list(qs.values("id", "name", "quantity_total", "quantity_used", "storage_location")[:3])
    return JsonResponse(data)

from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import UserPreference

@require_POST
def set_theme(request):
    theme = request.POST.get("theme")
    if theme not in {"light", "dark", "system"}:
        return HttpResponseBadRequest("invalid theme")
    resp = JsonResponse({"ok": True, "theme": theme})
    resp.set_cookie("theme", theme, max_age=60*60*24*365, samesite="Lax")
    if request.user.is_authenticated:
        prefs, _ = UserPreference.objects.get_or_create(user=request.user)
        prefs.theme = theme
        prefs.save(update_fields=["theme"])
    return resp



from django.db.models import Q
from django.utils.crypto import get_random_string
from django.urls import reverse
from django.http import HttpResponseForbidden

from .models import InventoryShare, InventoryItem

@login_required
def create_share_link(request):
    """Create or refresh a single active share link for the current user."""
    if request.method != "POST":
        return HttpResponseForbidden("POST required")
    # one active link per user
    token = get_random_string(32)
    share, _ = InventoryShare.objects.update_or_create(
        user=request.user, is_active=True,
        defaults={"token": token}
    )
    return redirect(reverse("inventory:settings") + "?tab=sharing")

@login_required
def revoke_share_link(request):
    if request.method != "POST":
        return HttpResponseForbidden("POST required")
    InventoryShare.objects.filter(user=request.user, is_active=True).update(is_active=False)
    return redirect(reverse("inventory:settings") + "?tab=sharing")

def shared_inventory(request, token):
    """Public, read-only inventory view by share token."""
    share = get_object_or_404(InventoryShare, token=token, is_active=True)
    owner = share.user

    q = (request.GET.get("q") or "").strip()
    cur_sort = (request.GET.get("sort") or "name").lower()
    cur_dir  = (request.GET.get("dir") or "asc").lower()

    items = InventoryItem.objects.filter(user=owner)

    if q:
        items = items.filter(
            Q(name__icontains=q) |
            Q(part_id__icontains=q) |
            Q(color__icontains=q) |
            Q(storage_location__icontains=q)
        )

    # available = total - used (stored field already? If you have a property, remove this)
    items = items.annotate(
        quantity_available_calc=ExpressionWrapper(
            F("quantity_total") - F("quantity_used"),
            output_field=IntegerField()
        )
    )

    sort_map = {
        "name":  "name",
        "part":  "part_id",
        "color": "color",
        "total": "quantity_total",
        "used":  "quantity_used",
        "avail": "quantity_available_calc",
        "loc":   "storage_location",
    }
    sort_field = sort_map.get(cur_sort, "name")
    order = sort_field if cur_dir != "desc" else f"-{sort_field}"
    items = items.order_by(order, "id")

    cfg = get_effective_config()
    per_page_public = getattr(cfg, "items_per_page", 50) or 50
    paginator = Paginator(items, per_page_public)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory/shared_inventory.html", {
        "owner": owner,
        "page_obj": page_obj,
        "items": page_obj.object_list,
        "q": q,
        "current_sort": cur_sort,
        "current_dir": cur_dir,
        "read_only": True,  # useful in template to hide edit/delete
    })
    
User = get_user_model()

def _get_owner_from_request(request):
    owner = request.user
    oid = request.GET.get("owner")
    if not oid:
        return owner
    try:
        candidate = User.objects.get(pk=oid)
    except User.DoesNotExist:
        raise Http404("Owner not found")

    if request.user.is_authenticated and (candidate == request.user):
        return candidate

    # Must be an accepted, active collaborator
    ok = InventoryCollab.objects.filter(
        owner=candidate,
        collaborator=request.user,
        is_active=True,
        accepted_at__isnull=False,
    ).exists()
    if not ok:
        raise Http404("Not allowed to view this inventory")
    return candidate

def _get_collab_record(viewer, owner):
    if not viewer.is_authenticated or viewer == owner:
        return None
    return InventoryCollab.objects.filter(
        owner=owner, collaborator=viewer, is_active=True, accepted_at__isnull=False
    ).first()

def _can_edit(viewer, owner):
    if viewer == owner:
        return True
    rel = _get_collab_record(viewer, owner)
    return bool(rel and rel.can_edit)

def _can_delete(viewer, owner):
    if viewer == owner:
        return True
    rel = _get_collab_record(viewer, owner)
    return bool(rel and rel.can_delete)
    
def error_404(request, exception, template_name="errors/404.html"):
    # Keep the request context so {% if request.user.is_authenticated %} works
    return render(request, template_name, status=404)

def error_500(request, template_name="errors/500.html"):
    return render(request, template_name, status=500)
    
def get_appcfg():
    return AppConfig.get_cached()
